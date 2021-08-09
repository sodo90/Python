import tkinter
from tkinter import *
from tkinter import ttk
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment


class LeadManGUI:

    def __init__(self, root):
        root.title("Lead Manager")
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        root.resizable(False, False)
        self.frame = ttk.Frame(root, padding='3 3 3 3')
        self.frame.grid(column=0, row=0, sticky=(N, S, E, W))
        self.prevfocus = ''
        self.index = StringVar()
        self.location = StringVar()
        self.businessname = StringVar()
        self.businessaddress = StringVar()
        self.contactname = StringVar()
        self.contactphone = StringVar()
        self.initialcontact = StringVar()
        self.method = StringVar()
        self.leadstatus = StringVar()
        self.leadnotes = StringVar()
        self.suggestedfollowup = StringVar()
        self.response = StringVar()
        self.lead_notes = tkinter.Text(self.frame, wrap=WORD)
        self.actiontrack = ttk.Treeview(self.frame, selectmode="browse", columns=['#1', '#2'], show='headings')
        self.actiontrack.bind("<ButtonRelease-1>", self.mousebind)
        self.buildwindow()

    def buildwindow(self):
        self.actiontrack.heading('#1', text='ID')
        self.actiontrack.column('#1', width=25)
        self.actiontrack.heading('#2', text='Customer Leads')
        eid = 1
        for lead in leadList:
            try:
                self.actiontrack.insert('', END, values=(eid, lead.bname), iid=eid)
                eid += 1
            except IndexError as e:
                print(e)
                pass
            except Exception as e:
                print(e)
                pass
        self.actiontrack.grid(column=0, row=0, rowspan=15, sticky=(N, S, E, W))
        self.actiontrack.focus(self.actiontrack.get_children()[0])
        self.actiontrack.selection_set(self.actiontrack.get_children()[0])
        self.prevfocus = self.actiontrack.focus()
        self.selectlead()

        newlead = ttk.Button(self.frame, command=self.newlead, text="New Lead")
        newlead.grid(column=0, row=15, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Location').grid(column=1, row=0, sticky=(N, S, E, W))
        location_entry = ttk.Entry(self.frame, textvariable=self.location)
        location_entry.grid(column=1, row=1, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Business Name').grid(column=2, row=0, sticky=(N, S, E, W))
        business_entry = ttk.Entry(self.frame, textvariable=self.businessname)
        business_entry.grid(column=2, row=1, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Business Address').grid(column=3, row=0, sticky=(N, S, E, W))
        business_address = ttk.Entry(self.frame, textvariable=self.businessaddress)
        business_address.grid(column=3, row=1, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Contact Name').grid(column=1, row=2, sticky=(N, S, E, W))
        contact_entry = ttk.Entry(self.frame, textvariable=self.contactname)
        contact_entry.grid(column=1, row=3, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Contact Number').grid(column=2, row=2, sticky=(N, S, E, W))
        contact_phone = ttk.Entry(self.frame, textvariable=self.contactphone)
        contact_phone.grid(column=2, row=3, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Initial Contact Date').grid(column=3, row=2, sticky=(N, S, E, W))
        initial_contact = ttk.Entry(self.frame, textvariable=self.initialcontact)
        initial_contact.grid(column=3, row=3, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Contact Method').grid(column=1, row=4, sticky=(N, S, E, W))
        method_entry = ttk.Entry(self.frame, textvariable=self.method)
        method_entry.grid(column=1, row=5, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Customer Possibility').grid(column=2, row=4, sticky=(N, S, E, W))
        lead_status = ttk.Combobox(self.frame,
                                   values=['Potential', 'Unlikely', 'Unknown'],
                                   textvariable=self.leadstatus)
        lead_status.grid(column=2, row=5, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Follow-up').grid(column=3, row=4, sticky=(N, S, E, W))
        suggested_followup = ttk.Entry(self.frame, textvariable=self.suggestedfollowup)
        suggested_followup.grid(column=3, row=5, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Response').grid(column=1, row=6, sticky=(N, S, E, W))
        response_entry = ttk.Entry(self.frame, textvariable=self.response)
        response_entry.grid(column=1, row=7, columnspan=2, sticky=(N, S, E, W))

        ttk.Label(self.frame, text='Notes and Comments').grid(column=2, row=9, sticky=(N, S))
        self.lead_notes.grid(column=1, row=10, columnspan=3, rowspan=5, sticky=(N, S, E, W))

        time_stamp = ttk.Button(self.frame, text='Note Timestamp', command=self.timestamp)
        time_stamp.grid(column=1, row=15, rowspan=2, sticky=(N, S, E, W))

        update_lead = ttk.Button(self.frame, text='Update Lead', command=self.updatelead)
        update_lead.grid(column=2, row=15, rowspan=2, sticky=(N, S, E, W))

        print_report = ttk.Button(self.frame, text='Print Report *Not Implemented*', command=self.printreport)
        print_report.grid(column=3, row=15, rowspan=2, sticky=(N, S, E, W))

    def mousebind(self, event):
        self.updatelead()
        self.selectlead()

    def selectlead(self):
        lead_object = leadList[int(self.actiontrack.focus())-1]
        self.index.set(lead_object.index)
        self.location.set(lead_object.location)
        self.businessname.set(lead_object.bname)
        self.businessaddress.set(lead_object.address)
        self.contactname.set(lead_object.contact)
        self.contactphone.set(lead_object.phone)
        self.initialcontact.set(lead_object.contactdate)
        self.method.set(lead_object.method)
        self.leadstatus.set(lead_object.leadstatus)
        self.leadnotes.set(lead_object.comments)
        self.lead_notes.delete('1.0', END)
        self.lead_notes.insert(END, self.leadnotes.get())
        self.suggestedfollowup.set(lead_object.followup)
        self.response.set(lead_object.response)


    def removelead(self):
        pass
        # TODO: Implement ability to remove a lead from the lead database

    def newlead(self):
        leadList.append(NewLead(str(len(leadList)+1), '', 'New Lead', '', '', '', '', '', '', '', '', ''))
        try:
            self.actiontrack.insert('', END, values=(leadList[-1].index, leadList[-1].bname), iid=leadList[-1].index)
        except IndexError as e:
            print(e)
            pass
        except Exception as e:
            print(e)
            pass
        self.actiontrack.focus(self.actiontrack.get_children()[-1])
        self.actiontrack.selection_set(self.actiontrack.get_children()[-1])
        self.selectlead()

    def timestamp(self):
        now = datetime.now()
        if len(self.lead_notes.get('1.0', END)) > 1:
            self.lead_notes.insert(END, f"\n\n{now.strftime('%B %d %Y @ %H:%M - ')}")
        else:
            self.lead_notes.insert(END, f"{now.strftime('%B %d %Y @ %H:%M - ')}")

    def updatelead(self):
        try:
            target = leadList[int(self.prevfocus)-1]
            target.location = self.location.get()
            target.bname = self.businessname.get()
            target.address = self.businessaddress.get()
            target.contact = self.contactname.get()
            target.phone = self.contactphone.get()
            target.contactdate = self.initialcontact.get()
            target.method = self.method.get()
            target.leadstatus = self.leadstatus.get()
            target.comments = self.lead_notes.get('1.0', 'end-1c')
            target.followup = self.suggestedfollowup.get()
            target.response = self.response.get()
            self.actiontrack.item(int(self.prevfocus), values=(target.index, target.bname))
            excelrow = int(target.index)+3
            # index
            ws.cell(row=excelrow, column=1, value=target.index).border = thin_border
            ws.cell(row=excelrow, column=1).alignment = centerx_centery
            # location
            ws.cell(row=excelrow, column=2, value=target.location).border = thin_border
            ws.cell(row=excelrow, column=2).alignment = centerx_centery
            # business name
            ws.cell(row=excelrow, column=3, value=target.bname).border = thin_border
            ws.cell(row=excelrow, column=3). alignment = centerx_centery
            # business address
            ws.cell(row=excelrow, column=4, value=target.address).border = thin_border
            ws.cell(row=excelrow, column=4).alignment = centerx_centery
            # business phone number
            ws.cell(row=excelrow, column=5, value=target.phone).border = thin_border
            ws.cell(row=excelrow, column=5).alignment = centerx_centery
            # contact names
            ws.cell(row=excelrow, column=6, value=target.contact).border = thin_border
            ws.cell(row=excelrow, column=6).alignment = centerx_centery
            # initial contact date
            ws.cell(row=excelrow, column=7, value=target.contactdate).border = thin_border
            ws.cell(row=excelrow, column=7).alignment = centerx_centery
            # method of initial contact
            ws.cell(row=excelrow, column=8, value=target.method).border = thin_border
            ws.cell(row=excelrow, column=8).alignment = centerx_centery
            # lead status
            ws.cell(row=excelrow, column=9, value=target.leadstatus).border = thin_border
            ws.cell(row=excelrow, column=9).alignment = centerx_centery
            # notes and comments
            ws.cell(row=excelrow, column=10, value=target.comments).border = thin_border
            ws.cell(row=excelrow, column=10).alignment = centerx_centery
            # follow-up status
            ws.cell(row=excelrow, column=11, value=target.followup).border = thin_border
            ws.cell(row=excelrow, column=11).alignment = centerx_centery
            # last response from customer
            ws.cell(row=excelrow, column=12, value=target.response).border = thin_border
            ws.cell(row=excelrow, column=12).alignment = centerx_centery
            self.prevfocus = str(self.actiontrack.focus())
            wb.save(r'bgit.xlsx')
        except Exception as e:
            pass

    def printreport(self):
        pass
        # TODO: Implement PDF report printing


class NewLead:
    def __init__(self, index, location, bname, address, phone, contact, contactdate,
                 method, leadstatus, comments, followup, response):
        self.index = index
        self.location = location
        self.bname = bname
        self.address = address
        self.contact = contact
        self.phone = phone
        self.contactdate = contactdate
        self.method = method
        self.leadstatus = leadstatus
        self.comments = comments
        self.followup = followup
        self.response = response


# Entry into the main program loop
leadList = []
nextlead = 0
centerx_centery = Alignment(horizontal='center',
                            vertical='center',
                            wrap_text=True)
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin')
                     )

wb = load_workbook('bgit.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=4, values_only=True):
    leadList.append(NewLead(str(nextlead+1), row[1], row[2], row[3], row[4], row[5],
                            row[6], row[7], row[8], row[9], row[10], row[11]))
    nextlead += 1
mainwindow = Tk()
LeadManGUI(mainwindow)
mainwindow.mainloop()
